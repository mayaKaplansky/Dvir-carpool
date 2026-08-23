"""
Regenerates the schedule data embedded inside index.html, from the Excel file.

Run this every time the Excel file is edited, then redeploy (see README.md).

Requires: pip install openpyxl

Usage:
    python update_data.py "path/to/הסעות דבירה תשפז - לוח משובץ.xlsx"
"""
import sys
import re
import json
import openpyxl

ROSTER_SHEET = "רשימת ילדים והורים"
SCHEDULE_SHEET = "לוח הסעות מוצע"
COUNT_SHEET = "ספירת הסעות למשפחה"


def norm(x):
    return str(x).strip() if x is not None else ""


def build_phone_lookup(wb):
    """
    Re-derives each family's phone numbers from the roster tab, matching
    families the same way the schedule was originally built: group student
    rows by (father name, mother name), tolerating a phone number that's
    filled in on one sibling's row but missing on another's.
    """
    ws = wb[ROSTER_SHEET]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    # columns: שם תלמיד, כיתת אם, ישוב, אב, נייד, אם, נייד 2
    students = []
    for r in rows:
        if r[0] is None:
            continue
        students.append({
            "name": norm(r[0]), "father": norm(r[3]), "father_mobile": norm(r[4]),
            "mother": norm(r[5]), "mother_mobile": norm(r[6]),
        })

    name_groups = {}
    for s in students:
        name_groups.setdefault((s["father"], s["mother"]), []).append(s)

    def common_prefix(names):
        token_lists = [n.split() for n in names]
        minlen = min(len(t) for t in token_lists)
        prefix = []
        for i in range(minlen):
            toks = set(t[i] for t in token_lists)
            if len(toks) == 1:
                prefix.append(token_lists[0][i])
            else:
                break
        return prefix

    phones = {}
    for (father, mother), members in name_groups.items():
        names = [m["name"] for m in members]
        if len(members) > 1:
            prefix = common_prefix(names)
            last_name = " ".join(prefix) if prefix else names[0].split()[0]
        else:
            toks = names[0].split()
            last_name = toks[0] if len(toks) == 1 else " ".join(toks[:-1])

        if father and mother:
            display = f"{last_name} - {father}/{mother}"
        elif father:
            display = f"{last_name} - {father}"
        elif mother:
            display = f"{last_name} - {mother}"
        else:
            display = last_name

        fmob = next((m["father_mobile"] for m in members if m["father_mobile"]), "")
        mmob = next((m["mother_mobile"] for m in members if m["mother_mobile"]), "")
        phones[display] = " / ".join(p for p in [fmob, mmob] if p)

    return phones


def main():
    if len(sys.argv) != 2:
        print("Usage: python update_data.py <path-to-xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    # data_only=True reads the last-saved calculated values. If you edited the
    # file in Excel/Google Sheets/LibreOffice and saved it, those values are
    # already there. If you edited it with a tool that doesn't calculate
    # formulas (e.g. wrote cells with openpyxl directly), open and save the
    # file once in Excel/LibreOffice first so the formula results are cached.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb[SCHEDULE_SHEET]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    schedule = []
    for date, shift, b1, b2, b3, car in rows:
        if date is None:
            continue
        schedule.append({
            "date": date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date),
            "shift": shift,
            "bus1": b1,
            "bus2": b2,
            "bus3": b3,
            "car": car,
        })

    ws3 = wb[COUNT_SHEET]
    fam_rows = list(ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True))
    families = sorted(set(r[0] for r in fam_rows if r[0]))

    phones = build_phone_lookup(wb)
    # only keep phone entries relevant to current families (keeps the payload tidy)
    phones = {name: phones.get(name, "") for name in families}

    unmatched = [name for name, p in phones.items() if p == "" and name not in phones]
    data = {"schedule": schedule, "families": families, "phones": phones}
    json_str = json.dumps(data, ensure_ascii=False, indent=2).replace("</", "<\\/")

    with open("index.html", encoding="utf-8") as f:
        html = f.read()

    pattern = re.compile(
        r'(<script id="schedule-data" type="application/json">\n)(.*?)(\n</script>)',
        re.S
    )
    if not pattern.search(html):
        print("Could not find the embedded data block in index.html -- "
              "make sure you're running this in the same folder as the "
              "original index.html.")
        sys.exit(1)

    html = pattern.sub(lambda m: m.group(1) + json_str + m.group(3), html, count=1)

    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)

    no_phone = sum(1 for p in phones.values() if not p)
    print(f"Updated index.html: {len(schedule)} schedule rows, {len(families)} families "
          f"({no_phone} with no phone number on file)")

    missing = [name for name in families if name not in phones or not phones[name]]
    if missing:
        print("No phone found for:", ", ".join(missing))


if __name__ == "__main__":
    main()
